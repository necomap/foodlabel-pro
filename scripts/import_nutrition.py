#!/usr/bin/env python3
# ============================================================
# scripts/import_nutrition.py
# 日本食品標準成分表2020年版のデータをPostgreSQLにインポート
#
# データソース:
#   文部科学省 食品成分データベース
#   https://www.mext.go.jp/a_menu/syokuhinseibun/mext_00001.html
#
# 使い方:
#   # 1. 文科省サイトからExcelをダウンロード（日本食品標準成分表2020年版）
#   # 2. 以下を実行
#   python3 scripts/import_nutrition.py \
#       --file /path/to/食品成分表.xlsx \
#       --db "postgresql://user:pass@localhost/foodlabel_pro"
#
# または既存のExcel(_resipi.xlsm)の「栄養成分表」シートからも取り込み可能
# ============================================================

import argparse
import sys
from pathlib import Path

try:
    import openpyxl
    import psycopg2
    import psycopg2.extras
except ImportError as e:
    print(f"エラー: {e}")
    print("pip install openpyxl psycopg2-binary を実行してください")
    sys.exit(1)


def safe_num(val):
    """数値変換（Tr=微量→0.001, (0)=0, -/NaN→None）"""
    if val is None:
        return None
    s = str(val).strip()
    if s in ('-', 'NaN', 'nan', '', '－'):
        return None
    if s in ('Tr', 'tr', 'TR'):
        return 0.001  # 微量
    if s.startswith('(') and s.endswith(')'):
        try:
            return float(s[1:-1])
        except ValueError:
            return None
    try:
        return float(s)
    except ValueError:
        return None


def import_from_mext_excel(filepath: str, conn):
    """文科省Excel形式からインポート"""
    print(f"ファイルを読み込み中: {filepath}")
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)

    # シートを探す
    sheet = None
    for name in wb.sheetnames:
        if '成分表' in name or 'Table' in name.lower():
            sheet = wb[name]
            break
    if not sheet:
        sheet = wb.active

    print(f"シート: {sheet.title}")
    rows = list(sheet.iter_rows(values_only=True))

    # ヘッダー行を検索（食品番号が含まれる行）
    header_row = -1
    for i, row in enumerate(rows[:10]):
        if row and any('食品番号' in str(c or '') or 'ENERC' in str(c or '') for c in row):
            header_row = i
            break

    if header_row < 0:
        print("警告: ヘッダーが見つかりません。5行目からデータとして処理します")
        header_row = 4

    headers = [str(c or '') for c in rows[header_row]]
    print(f"ヘッダー行: {header_row + 1}行目 / データ行数: {len(rows) - header_row - 1}行")

    def col(candidates):
        for name in candidates:
            for i, h in enumerate(headers):
                if name in h:
                    return i
        return -1

    # 列マッピング
    ci = {
        'id':     col(['食品番号', '0', 'No']),
        'group':  col(['食品群']),
        'name':   col(['食品名', '3']),
        'waste':  col(['廃棄率', '廃棄']),
        'energy': col(['エネルギー', 'ENERC_KCAL', 'kcal']),
        'water':  col(['水分', 'WATER']),
        'protein': col(['たんぱく質', 'PROT']),
        'fat':    col(['脂質', 'FAT']),
        'cholesterol': col(['コレステロール', 'CHOLE']),
        'carb':   col(['炭水化物', 'CHOCDF']),
        'fiber':  col(['食物繊維総量', '食物繊維', 'FIBTG']),
        'sugar':  col(['糖質']),
        'sodium': col(['ナトリウム', 'NA']),
        'salt':   col(['食塩相当量', 'NACL']),
    }
    print(f"列マッピング: {ci}")

    cur = conn.cursor()
    imported = 0
    errors   = []

    for row in rows[header_row + 1:]:
        if not row:
            continue
        # 食品番号を取得
        raw_id = row[ci['id']] if ci['id'] >= 0 else None
        if raw_id is None:
            continue
        try:
            food_id = int(str(raw_id).replace('-', '').strip())
        except (ValueError, TypeError):
            continue
        if food_id <= 0:
            continue

        name = str(row[ci['name']] or '').strip() if ci['name'] >= 0 else ''
        if not name:
            continue

        def get(key):
            idx = ci.get(key, -1)
            if idx < 0 or idx >= len(row):
                return None
            return safe_num(row[idx])

        try:
            cur.execute("""
                INSERT INTO nutrition_data (
                    id, food_group, food_name, waste_ratio,
                    energy_kcal, water, protein, fat, cholesterol,
                    carbohydrate, dietary_fiber, sugar, sodium, salt_equivalent,
                    data_version
                ) VALUES (
                    %s, %s, %s, %s,
                    %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s,
                    '2020'
                )
                ON CONFLICT (id) DO UPDATE SET
                    food_name      = EXCLUDED.food_name,
                    energy_kcal    = EXCLUDED.energy_kcal,
                    protein        = EXCLUDED.protein,
                    fat            = EXCLUDED.fat,
                    carbohydrate   = EXCLUDED.carbohydrate,
                    sodium         = EXCLUDED.sodium,
                    salt_equivalent = EXCLUDED.salt_equivalent,
                    dietary_fiber  = EXCLUDED.dietary_fiber,
                    data_version   = '2020'
            """, (
                food_id,
                str(row[ci['group']] or '').strip() if ci['group'] >= 0 else '',
                name,
                get('waste'),
                get('energy'), get('water'), get('protein'), get('fat'), get('cholesterol'),
                get('carb'), get('fiber'), get('sugar'), get('sodium'), get('salt'),
            ))
            imported += 1
        except Exception as e:
            errors.append(f"ID {food_id} ({name}): {e}")
            conn.rollback()
            continue

    conn.commit()
    cur.close()
    print(f"\n結果: {imported}件インポート, {len(errors)}件エラー")
    if errors:
        for err in errors[:10]:
            print(f"  エラー: {err}")
    return imported


def main():
    parser = argparse.ArgumentParser(description='日本食品標準成分表インポート')
    parser.add_argument('--file', required=True, help='Excelファイルのパス')
    parser.add_argument('--db',   required=True, help='PostgreSQL接続文字列')
    args = parser.parse_args()

    if not Path(args.file).exists():
        print(f"エラー: ファイルが見つかりません: {args.file}")
        sys.exit(1)

    print("=" * 50)
    print("日本食品標準成分表 インポートスクリプト")
    print("=" * 50)

    conn = psycopg2.connect(args.db)
    print("データベース接続成功")

    try:
        imported = import_from_mext_excel(args.file, conn)
        print(f"\n✅ 完了: {imported}件の食品データをインポートしました")
    finally:
        conn.close()


if __name__ == '__main__':
    main()
