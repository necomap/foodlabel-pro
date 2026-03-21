#!/usr/bin/env python3
# ============================================================
# scripts/migration_from_excel.py
# 既存Excelファイル（_resipi.xlsm）からPostgreSQLへの移行スクリプト
#
# 必要パッケージ: pip install openpyxl psycopg2-binary python-dotenv
#
# 使い方:
#   python3 scripts/migration_from_excel.py \
#       --file /path/to/_resipi.xlsm \
#       --db "postgresql://user:pass@localhost/foodlabel_pro" \
#       --user-id "YOUR_USER_UUID"
# ============================================================

import argparse
import json
import sys
import csv
import unicodedata
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("エラー: openpyxlが必要です。pip install openpyxl を実行してください")
    sys.exit(1)

try:
    import psycopg2
    import psycopg2.extras
except ImportError:
    print("エラー: psycopg2が必要です。pip install psycopg2-binary を実行してください")
    sys.exit(1)


# ============================================================
# ユーティリティ
# ============================================================

def to_fullwidth(s: str) -> str:
    """半角文字を全角に変換する"""
    if not s:
        return s
    result = ""
    for c in s:
        cp = ord(c)
        if 0x21 <= cp <= 0x7E:  # 半角英数字・記号
            result += chr(cp + 0xFEE0)
        elif c == ' ':
            result += '　'
        else:
            result += c
    return result


def normalize_name(s: str) -> str:
    """食材名・品名を正規化する（全角統一、NFKC正規化）"""
    if not s:
        return ""
    s = str(s).strip()
    s = unicodedata.normalize('NFKC', s)  # 全角カタカナ → 半角カナを全角に
    return s


def safe_float(val) -> float | None:
    """数値に変換（失敗したらNone）"""
    if val is None or str(val).strip() in ('', 'NaN', 'nan', '-'):
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def safe_int(val) -> int | None:
    """整数に変換（失敗したらNone）"""
    f = safe_float(val)
    return int(f) if f is not None else None


# ============================================================
# Excelパーサー
# ============================================================

class ExcelParser:
    def __init__(self, filepath: str):
        print(f"Excelファイルを読み込んでいます: {filepath}")
        self.wb = openpyxl.load_workbook(filepath, data_only=True)
        print(f"シート一覧: {self.wb.sheetnames}")

    def get_recipes(self) -> list[dict]:
        """DBシートからレシピデータを取得する"""
        # DBシートを探す
        sheet_name = None
        for name in self.wb.sheetnames:
            if name == 'DB':
                sheet_name = name
                break
        if not sheet_name:
            for name in self.wb.sheetnames:
                if name.startswith('DB'):
                    sheet_name = name
                    break

        if not sheet_name:
            print(f"警告: DBシートが見つかりません（使用可能: {self.wb.sheetnames}）")
            return []

        ws = self.wb[sheet_name]
        print(f"シート '{sheet_name}' を読み込みます（{ws.max_row}行 × {ws.max_column}列）")

        # ヘッダー行を探す
        headers = []
        header_row = 1
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row and (row[0] == 'No' or row[3] == '品名' or row[2] == '品名'):
                headers = [str(c) if c is not None else '' for c in row]
                header_row = row_idx
                break
            if row_idx > 5:
                break

        if not headers:
            # ヘッダーなし：デフォルト列順を仮定
            print("警告: ヘッダー行が見つかりません。デフォルト列順を使用します")
            header_row = 1

        def col(name: str, fallback: int = -1) -> int:
            """ヘッダー名から列インデックスを取得"""
            if not headers:
                return fallback
            try:
                return next(i for i, h in enumerate(headers) if name in (h or ''))
            except StopIteration:
                return fallback

        # 列インデックス
        ci = {
            'no':       col('No', 0),
            'category': col('カテゴリ', 1),
            'name':     col('品名', 3),
            'kana':     col('カナ', 4),
            'unit_cnt': col('仕上数量', 5),
            'sale_price': col('販売価格', 9),
            'cost_rate': col('原価率', 10),
            'shelf':    col('賞味期限', 11),
            'ingreds':  col('原材料', 12),
            'notes':    col('注意事項', 13),
            'steam1':   col('スチーム1', -1),
            'top1':     col('上火1', -1),
            'bot1':     col('下火1', -1),
            'time1':    col('時間1', -1),
            'steam2':   col('スチーム2', -1),
            'top2':     col('上火2', -1),
            'bot2':     col('下火2', -1),
            'time2':    col('時間2', -1),
            'mat1':     col('材料1', -1),
            'step1':    col('手順1', -1),
        }

        recipes = []
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row_idx <= header_row:
                continue
            if not row or not row[ci['name'] if ci['name'] > -1 else 3]:
                continue

            name_raw = row[ci['name'] if ci['name'] > -1 else 3]
            name = normalize_name(str(name_raw)).strip()
            if not name:
                continue

            # 材料を抽出（最大30種）
            ingredients = []
            mat_start = ci['mat1']
            if mat_start > -1:
                for m in range(30):
                    base = mat_start + m * 14  # 材料Nの列幅（名前・分量・単位・順位・原価 + 栄養9列）
                    if base >= len(row):
                        break
                    mat_name = row[base] if base < len(row) else None
                    if not mat_name:
                        continue
                    mat_name = normalize_name(str(mat_name)).strip()
                    if not mat_name:
                        continue

                    amount_val = row[base + 1] if base + 1 < len(row) else None
                    unit_val   = row[base + 2] if base + 2 < len(row) else 'g'
                    order_val  = row[base + 3] if base + 3 < len(row) else m
                    cost_val   = row[base + 4] if base + 4 < len(row) else None

                    ingredients.append({
                        'name':   mat_name,
                        'amount': safe_float(amount_val) or 0,
                        'unit':   str(unit_val or 'g').strip() or 'g',
                        'order':  safe_int(order_val) or m,
                        'cost':   safe_float(cost_val),
                    })

            # 作り方を抽出（最大35手順）
            steps = []
            step_start = ci['step1']
            if step_start > -1:
                for s in range(35):
                    idx = step_start + s
                    if idx >= len(row):
                        break
                    step_text = row[idx]
                    if step_text:
                        steps.append(str(step_text).strip())

            # 焼成条件
            baking = []
            def mk_baking(steam_idx, top_idx, bot_idx, time_idx):
                if steam_idx < 0 or steam_idx >= len(row):
                    return None
                steam = row[steam_idx]
                top   = row[top_idx]   if top_idx  > -1 and top_idx  < len(row) else None
                bot   = row[bot_idx]   if bot_idx  > -1 and bot_idx  < len(row) else None
                time  = row[time_idx]  if time_idx > -1 and time_idx < len(row) else None
                if not any([steam, top, bot, time]):
                    return None
                return {
                    'steam':      str(steam or '').upper() or None,
                    'topHeat':    safe_float(top),
                    'bottomHeat': safe_float(bot),
                    'timeMin':    safe_float(time),
                }
            b1 = mk_baking(ci['steam1'], ci['top1'], ci['bot1'], ci['time1'])
            b2 = mk_baking(ci['steam2'], ci['top2'], ci['bot2'], ci['time2'])
            if b1: baking.append(b1)
            if b2: baking.append(b2)

            recipes.append({
                'no':            safe_int(row[ci['no']]) or row_idx,
                'category':      normalize_name(str(row[ci['category']] or '')).strip(),
                'name':          name,
                'name_kana':     normalize_name(str(row[ci['kana']] or '')).strip(),
                'unit_count':    safe_int(row[ci['unit_cnt']]) or 1,
                'sale_price':    safe_float(row[ci['sale_price']]),
                'shelf_life_days': safe_int(row[ci['shelf']]),
                'ingredients':   ingredients,
                'steps':         steps,
                'baking_conditions': baking,
                'notes':         str(row[ci['notes']] or '').strip(),
            })

        print(f"レシピ {len(recipes)} 件を解析しました")
        return recipes

    def get_nutrition_data(self) -> list[dict]:
        """栄養成分表シートからデータを取得する"""
        sheet_name = '栄養成分表'
        if sheet_name not in self.wb.sheetnames:
            print(f"警告: '{sheet_name}' シートが見つかりません")
            return []

        ws = self.wb[sheet_name]
        print(f"栄養成分表シートを読み込みます")
        
        rows = list(ws.iter_rows(values_only=True))
        # ヘッダーを探す（食品番号、食品名が含まれる行）
        header_row = -1
        for i, row in enumerate(rows):
            if row and any('食品名' in str(c or '') for c in row):
                header_row = i
                break

        if header_row < 0:
            print("警告: 栄養成分表のヘッダーが見つかりません")
            return []

        headers = [str(c or '') for c in rows[header_row]]
        
        def col(name):
            for i, h in enumerate(headers):
                if name in h:
                    return i
            return -1

        ci_n = {
            'id':            col('食品番号'),
            'group':         col('食品群'),
            'name':          col('食品名'),
            'waste':         col('廃棄'),
            'energy':        col('エネルギー（kcal）'),
            'water':         col('水分'),
            'protein':       col('たんぱく質'),
            'fat':           col('脂質'),
            'cholesterol':   col('コレステロール'),
            'carbohydrate':  col('炭水化物'),
            'fiber_sol':     col('水溶性食物繊維'),
            'fiber_insol':   col('不溶性食物繊維'),
            'fiber':         col('食物繊維総量'),
            'sugar':         col('糖質'),
            'ash':           col('灰分'),
            'sodium':        col('ナトリウム'),
            'salt':          col('食塩相当量'),
        }

        data = []
        for row in rows[header_row + 1:]:
            if not row or not row[ci_n['id']] or not row[ci_n['name']]:
                continue
            food_id = safe_int(row[ci_n['id']])
            if not food_id:
                continue

            data.append({
                'id':            food_id,
                'food_group':    str(row[ci_n['group']] or '').strip(),
                'food_name':     str(row[ci_n['name']] or '').strip(),
                'waste_ratio':   safe_float(row[ci_n['waste']]),
                'energy_kcal':   safe_float(row[ci_n['energy']]),
                'water':         safe_float(row[ci_n['water']]),
                'protein':       safe_float(row[ci_n['protein']]),
                'fat':           safe_float(row[ci_n['fat']]),
                'cholesterol':   safe_float(row[ci_n['cholesterol']]),
                'carbohydrate':  safe_float(row[ci_n['carbohydrate']]),
                'dietary_fiber': safe_float(row[ci_n['fiber']]),
                'sugar':         safe_float(row[ci_n['sugar']]),
                'sodium':        safe_float(row[ci_n['sodium']]),
                'salt_equivalent': safe_float(row[ci_n['salt']]),
            })

        print(f"栄養成分 {len(data)} 件を解析しました")
        return data


# ============================================================
# DBインサーター
# ============================================================

class DBInserter:
    def __init__(self, db_url: str):
        self.conn = psycopg2.connect(db_url)
        self.cur  = self.conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        print("データベースに接続しました")

    def upsert_nutrition_data(self, data: list[dict]) -> None:
        """栄養成分表データをupsert"""
        if not data:
            return
        print(f"栄養成分表を登録中 ({len(data)}件)...")
        for row in data:
            self.cur.execute("""
                INSERT INTO nutrition_data (
                    id, food_group, food_name, waste_ratio,
                    energy_kcal, water, protein, fat, cholesterol,
                    carbohydrate, dietary_fiber, sugar, sodium, salt_equivalent,
                    data_version
                ) VALUES (
                    %(id)s, %(food_group)s, %(food_name)s, %(waste_ratio)s,
                    %(energy_kcal)s, %(water)s, %(protein)s, %(fat)s, %(cholesterol)s,
                    %(carbohydrate)s, %(dietary_fiber)s, %(sugar)s, %(sodium)s, %(salt_equivalent)s,
                    '2015'
                )
                ON CONFLICT (id) DO UPDATE SET
                    food_name = EXCLUDED.food_name,
                    energy_kcal = EXCLUDED.energy_kcal,
                    protein = EXCLUDED.protein,
                    fat = EXCLUDED.fat,
                    carbohydrate = EXCLUDED.carbohydrate,
                    sodium = EXCLUDED.sodium,
                    salt_equivalent = EXCLUDED.salt_equivalent
            """, row)
        self.conn.commit()
        print(f"  → 栄養成分表 {len(data)}件 完了")

    def find_or_create_category(self, user_id: str, name: str) -> str | None:
        """カテゴリを検索または作成する"""
        if not name:
            return None
        self.cur.execute(
            "SELECT id FROM categories WHERE name = %s AND (user_id = %s OR user_id IS NULL) LIMIT 1",
            (name, user_id)
        )
        row = self.cur.fetchone()
        if row:
            return str(row['id'])

        self.cur.execute(
            "INSERT INTO categories (id, user_id, name) VALUES (gen_random_uuid(), %s, %s) RETURNING id",
            (user_id, name)
        )
        return str(self.cur.fetchone()['id'])

    def find_ingredient(self, user_id: str, name: str) -> str | None:
        """食材マスタを検索する"""
        self.cur.execute(
            "SELECT id FROM ingredients WHERE name = %s AND (user_id = %s OR user_id IS NULL) AND is_active = TRUE LIMIT 1",
            (name, user_id)
        )
        row = self.cur.fetchone()
        return str(row['id']) if row else None

    def create_ingredient(self, user_id: str, name: str) -> str:
        """食材マスタを新規作成する"""
        allergens = self._detect_allergens(name)
        self.cur.execute("""
            INSERT INTO ingredients (id, user_id, name, name_kana, allergens, is_active)
            VALUES (gen_random_uuid(), %s, %s, '', %s, TRUE)
            RETURNING id
        """, (user_id, name, allergens))
        return str(self.cur.fetchone()['id'])

    def _detect_allergens(self, name: str) -> list[str]:
        """食材名からアレルゲンを自動判定（簡易版）"""
        ALLERGEN_KEYWORDS = {
            '小麦': ['小麦', '強力粉', '薄力粉', '中力粉', '準強力粉', '全粒粉', 'ライ麦', 'ふすま'],
            '卵':   ['卵', '全卵', '卵黄', '卵白', 'たまご'],
            '乳':   ['牛乳', '生クリーム', 'バター', 'チーズ', 'ヨーグルト', 'クリームチーズ', 'マスカルポーネ'],
            'くるみ': ['くるみ', 'クルミ'],
            '大豆': ['豆乳', '大豆', '枝豆', '酒粕'],
            '落花生': ['ピーナッツ', '落花生'],
            'えび': ['えび', 'エビ'],
            'かに': ['かに', 'カニ'],
        }
        found = []
        for allergen, keywords in ALLERGEN_KEYWORDS.items():
            if any(kw in name for kw in keywords):
                found.append(allergen)
        return found

    def insert_recipes(self, recipes: list[dict], user_id: str, overwrite: bool = False) -> dict:
        """レシピをDBに登録する"""
        imported = 0
        skipped  = 0
        errors   = []

        for recipe in recipes:
            try:
                name = recipe['name']
                if not name:
                    skipped += 1
                    continue

                # 上書きしない場合はスキップ
                if not overwrite:
                    self.cur.execute(
                        "SELECT id FROM recipes WHERE user_id = %s AND name = %s AND is_active = TRUE",
                        (user_id, name)
                    )
                    if self.cur.fetchone():
                        skipped += 1
                        continue

                # 上書きの場合は既存を論理削除
                if overwrite:
                    self.cur.execute(
                        "UPDATE recipes SET is_active = FALSE WHERE user_id = %s AND name = %s",
                        (user_id, name)
                    )

                # カテゴリ
                cat_id = self.find_or_create_category(user_id, recipe.get('category', ''))

                # レシピ本体を挿入
                baking_json = json.dumps(recipe.get('baking_conditions', []), ensure_ascii=False) \
                    if recipe.get('baking_conditions') else None

                self.cur.execute("""
                    INSERT INTO recipes (
                        id, user_id, category_id, name, name_kana,
                        unit_count, shelf_life_days, sale_price, notes,
                        baking_conditions, is_active, created_at
                    ) VALUES (
                        gen_random_uuid(), %s, %s, %s, %s,
                        %s, %s, %s, %s,
                        %s::jsonb, TRUE, NOW()
                    ) RETURNING id
                """, (
                    user_id, cat_id, name, recipe.get('name_kana') or '',
                    recipe.get('unit_count', 1),
                    recipe.get('shelf_life_days'),
                    recipe.get('sale_price'),
                    recipe.get('notes') or '',
                    baking_json,
                ))
                recipe_id = str(self.cur.fetchone()['id'])

                # 材料を挿入
                for order_idx, ing in enumerate(recipe.get('ingredients', [])):
                    ing_name = ing.get('name', '').strip()
                    if not ing_name:
                        continue

                    # 食材マスタを検索（なければ作成）
                    ing_id = self.find_ingredient(user_id, ing_name)
                    if not ing_id:
                        ing_id = self.create_ingredient(user_id, ing_name)

                    self.cur.execute("""
                        INSERT INTO recipe_ingredients (
                            id, recipe_id, ingredient_id, amount, unit,
                            display_order, sort_by_weight, cost_price,
                            nutrition_unconfirmed
                        ) VALUES (
                            gen_random_uuid(), %s, %s, %s, %s,
                            %s, TRUE, %s,
                            TRUE
                        )
                    """, (
                        recipe_id, ing_id, ing.get('amount', 0), ing.get('unit', 'g'),
                        ing.get('order', order_idx),
                        ing.get('cost') / ing.get('amount') if (ing.get('cost') and ing.get('amount')) else None,
                    ))

                # 作り方を挿入
                for step_num, step_text in enumerate(recipe.get('steps', []), start=1):
                    if step_text.strip():
                        self.cur.execute("""
                            INSERT INTO recipe_steps (id, recipe_id, step_number, instruction)
                            VALUES (gen_random_uuid(), %s, %s, %s)
                        """, (recipe_id, step_num, step_text))

                self.conn.commit()
                imported += 1

            except Exception as e:
                self.conn.rollback()
                errors.append({'no': recipe.get('no', 0), 'name': recipe.get('name', ''), 'error': str(e)})
                print(f"  エラー: {recipe.get('name', '')} - {e}")

        return {'imported': imported, 'skipped': skipped, 'errors': errors}

    def close(self):
        self.cur.close()
        self.conn.close()


# ============================================================
# メイン処理
# ============================================================

def main():
    parser = argparse.ArgumentParser(description='FoodLabel Pro - Excelデータ移行スクリプト')
    parser.add_argument('--file',    required=True, help='Excelファイルのパス (_resipi.xlsm)')
    parser.add_argument('--db',      required=True, help='PostgreSQL接続文字列')
    parser.add_argument('--user-id', required=True, dest='user_id', help='インポート先ユーザーのUUID')
    parser.add_argument('--overwrite', action='store_true', help='既存レシピを上書きする')
    parser.add_argument('--nutrition-only', action='store_true', dest='nutrition_only',
                       help='栄養成分表のみをインポートする')
    parser.add_argument('--skip-nutrition', action='store_true', dest='skip_nutrition',
                       help='栄養成分表のインポートをスキップする')
    parser.add_argument('--log', default='migration_log.csv', help='エラーログのファイルパス')
    args = parser.parse_args()

    if not Path(args.file).exists():
        print(f"エラー: ファイルが見つかりません: {args.file}")
        sys.exit(1)

    print("\n" + "=" * 60)
    print("FoodLabel Pro - データ移行スクリプト")
    print("=" * 60)
    print(f"ファイル  : {args.file}")
    print(f"ユーザーID: {args.user_id}")
    print(f"上書き    : {'有効' if args.overwrite else '無効'}")
    print("=" * 60 + "\n")

    start_time = datetime.now()

    # Excel解析
    parser_obj = ExcelParser(args.file)

    # DB接続
    db = DBInserter(args.db)

    results = {}

    # 栄養成分表のインポート
    if not args.skip_nutrition:
        print("\n[1/2] 栄養成分表をインポートします...")
        nutrition_data = parser_obj.get_nutrition_data()
        if nutrition_data:
            db.upsert_nutrition_data(nutrition_data)

    # レシピのインポート
    if not args.nutrition_only:
        print("\n[2/2] レシピをインポートします...")
        recipes = parser_obj.get_recipes()
        if recipes:
            results = db.insert_recipes(recipes, args.user_id, args.overwrite)
            print(f"\n結果:")
            print(f"  成功: {results.get('imported', 0)}件")
            print(f"  スキップ: {results.get('skipped', 0)}件")
            print(f"  エラー: {len(results.get('errors', []))}件")

            # エラーログ出力
            errors = results.get('errors', [])
            if errors:
                print(f"\nエラーログを {args.log} に書き込みます...")
                with open(args.log, 'w', newline='', encoding='utf-8-sig') as f:
                    writer = csv.DictWriter(f, fieldnames=['no', 'name', 'error'])
                    writer.writeheader()
                    writer.writerows(errors)

    db.close()

    elapsed = (datetime.now() - start_time).total_seconds()
    print(f"\n処理時間: {elapsed:.1f}秒")
    print("\n移行が完了しました！")
    print("次のステップ:")
    print("  1. アプリにログイン: http://localhost:3000/auth/login")
    print("  2. レシピ一覧を確認: http://localhost:3000/dashboard/recipes")
    print("  3. 食材マスタに成分情報を設定してください")


if __name__ == '__main__':
    main()
